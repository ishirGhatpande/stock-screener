# excel_output.py
# Writes a formatted Excel workbook with four sheets of results.

import os
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXCEL_FILENAME, CHART_FILENAME, TICKERS

# --- Colour palette ---
_HEADER_FILL  = "1F4E79"   # dark navy
_ACCENT_FILL  = "BDD7EE"   # light blue
_GOLD_FILL    = "FFD700"   # max-Sharpe highlight
_RED_FILL     = "FF6347"   # min-vol highlight
_INTERP_FILL  = "505050"   # dark grey for interpretation headers
_DIVIDER_FILL = "DDDDDD"   # light grey separator

# Sector labels used in plain-English interpretation
_SECTOR_MAP = {
    "AAPL": "Technology",          "MSFT": "Technology",          "NVDA": "Technology",
    "JPM":  "Financials",          "BAC":  "Financials",          "GS":   "Financials",
    "XOM":  "Energy",              "CVX":  "Energy",              "SLB":  "Energy",
    "JNJ":  "Healthcare",          "UNH":  "Healthcare",          "PFE":  "Healthcare",
    "PG":   "Consumer Staples",    "KO":   "Consumer Staples",    "AMZN": "Consumer Discretionary",
}


def _header_style(cell, fill_hex=_HEADER_FILL):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, max_width)


# ---------------------------------------------------------------------------
# Sheet 1 helpers – portfolio block + interpretation
# ---------------------------------------------------------------------------

def _write_portfolio_block(ws, start_row, label, portfolio, tickers, fill):
    """Write section title + metrics + weights block. Returns next available row."""
    row = start_row

    # Section title
    title = ws.cell(row=row, column=1, value=label)
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=fill)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 24
    row += 1

    # Column headers
    for col_idx, h in enumerate(["Metric / Ticker", "Value"], 1):
        _header_style(ws.cell(row=row, column=col_idx))
        ws.cell(row=row, column=col_idx, value=h)
    row += 1

    # Summary metrics
    for metric, value in [
        ("Expected Return",       f"{portfolio['Return']*100:.2f}%"),
        ("Annualised Volatility", f"{portfolio['Volatility']*100:.2f}%"),
        ("Sharpe Ratio",          f"{portfolio['Sharpe']:.4f}"),
    ]:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Weights sub-header
    wh = ws.cell(row=row, column=1, value="Portfolio Weights")
    wh.font = Font(italic=True, bold=True)
    row += 1

    for ticker in tickers:
        if ticker in portfolio.index:
            ws.cell(row=row, column=1, value=ticker)
            ws.cell(row=row, column=2, value=f"{portfolio[ticker]*100:.2f}%")
            row += 1

    return row


def _write_interpretation(ws, start_row, portfolio, expected_returns, volatilities, tickers, is_max_sharpe):
    """
    Write a plain-English explanation of why the optimizer chose these weights.
    Returns next available row.
    """
    row = start_row

    # --- Section header ---
    hdr = ws.cell(row=row, column=1, value="  Portfolio Interpretation")
    hdr.font = Font(bold=True, size=11, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor=_INTERP_FILL)
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22
    row += 1

    # Helper: write a single wrapped-text row spanning columns A:D
    def _para(text, row_height=None, bold=False):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(bold=bold, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        if row_height:
            ws.row_dimensions[row].height = row_height
        row += 1

    # --- Objective ---
    if is_max_sharpe:
        _para(
            "Objective:  Maximise the Sharpe ratio — the annualised excess return earned per unit "
            "of volatility. A higher Sharpe means more return for each unit of risk taken. This "
            "portfolio sits at the most efficient point on the frontier: no other combination of "
            "these 15 stocks delivers more return per unit of risk.",
            row_height=58,
        )
    else:
        _para(
            "Objective:  Minimise total portfolio volatility regardless of return. This is the "
            "safest point on the efficient frontier — the portfolio that fluctuates the least "
            "day-to-day. Investors who prioritise capital preservation over maximum growth would "
            "prefer this portfolio to the max-Sharpe alternative.",
            row_height=58,
        )

    # --- How it was built ---
    _para(
        "How it was built:  The optimizer generated 10,000 random long-only weight combinations "
        "(Dirichlet distribution — guarantees all weights >= 0 and sum to exactly 1). It computed "
        "the annualised return, volatility and Sharpe ratio for each using 3 years of daily log "
        "returns and a risk-free rate of 4.3%. " +
        ("The portfolio with the highest Sharpe ratio was selected." if is_max_sharpe
         else "The portfolio with the lowest annualised standard deviation was selected."),
        row_height=64,
    )

    row += 1  # blank spacer

    # --- Largest allocations mini-table ---
    _para("Largest Allocations:", bold=True, row_height=16)

    # Mini-table column headers
    for col_idx, h in enumerate(["Ticker (Sector)", "Weight", "Indiv. Return", "Indiv. Vol"], 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor=_ACCENT_FILL)
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Top 5 holdings by weight
    weights = {t: portfolio[t] for t in tickers if t in portfolio.index}
    for ticker, weight in sorted(weights.items(), key=lambda x: x[1], reverse=True)[:5]:
        sector = _SECTOR_MAP.get(ticker, "")
        ann_ret = expected_returns.get(ticker, 0) * 100
        ann_vol = volatilities.get(ticker, 0) * 100
        ws.cell(row=row, column=1, value=f"{ticker}  ({sector})")
        ws.cell(row=row, column=2, value=f"{weight * 100:.1f}%")
        ws.cell(row=row, column=3, value=f"{ann_ret:.1f}%")
        ws.cell(row=row, column=4, value=f"{ann_vol:.1f}%")
        row += 1

    row += 1  # blank spacer

    # --- Why these weights ---
    if is_max_sharpe:
        _para(
            "Why these weights?  The optimizer overweights stocks that delivered strong "
            "risk-adjusted returns AND that move differently from each other. A stock earns a "
            "large allocation by being individually efficient (high Sharpe) and by 'zigging' when "
            "the rest of the portfolio 'zags'. Combining low-correlation assets reduces "
            "portfolio-level volatility below the weighted average of individual volatilities — "
            "this is the core mathematical benefit of diversification.",
            row_height=72,
        )
    else:
        _para(
            "Why these weights?  The optimizer favours stocks with lower individual volatility "
            "and, critically, low pairwise correlations. When assets do not move in lockstep, "
            "their daily gains and losses partially cancel each other out. The result is a "
            "portfolio whose volatility is meaningfully lower than the weighted average of each "
            "stock's individual volatility — diversification in its purest form.",
            row_height=64,
        )

    return row


# ---------------------------------------------------------------------------
# Sheet 1 – Optimal Portfolio
# ---------------------------------------------------------------------------

def _write_optimal_sheet(wb, optimal, expected_returns, volatilities, tickers):
    ws = wb.create_sheet("Optimal Portfolio")

    # Sheet title banner
    t = ws.cell(row=1, column=1, value="Optimal Portfolio Analysis")
    t.font = Font(bold=True, size=15, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 30

    row = 3  # leave row 2 as breathing room

    # ---- Max Sharpe ----
    row = _write_portfolio_block(ws, row, "Max Sharpe Ratio Portfolio", optimal["max_sharpe"], tickers, _GOLD_FILL)
    row += 1
    row = _write_interpretation(ws, row, optimal["max_sharpe"], expected_returns, volatilities, tickers, is_max_sharpe=True)

    # Visual separator between the two portfolios
    row += 1
    sep = ws.cell(row=row, column=1, value="")
    sep.fill = PatternFill("solid", fgColor=_DIVIDER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 8
    row += 2

    # ---- Min Volatility ----
    row = _write_portfolio_block(ws, row, "Min Volatility Portfolio", optimal["min_vol"], tickers, _RED_FILL)
    row += 1
    _write_interpretation(ws, row, optimal["min_vol"], expected_returns, volatilities, tickers, is_max_sharpe=False)

    # Fixed column widths (auto-width would be misled by wide merged interpretation cells)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ---------------------------------------------------------------------------
# Sheet 2 – Efficient Frontier Data
# ---------------------------------------------------------------------------

def _write_frontier_sheet(wb, simulation):
    ws = wb.create_sheet("Efficient Frontier Data")

    display = simulation.copy()
    display["Return"]     = display["Return"].map(lambda x: f"{x*100:.2f}%")
    display["Volatility"] = display["Volatility"].map(lambda x: f"{x*100:.2f}%")
    display["Sharpe"]     = display["Sharpe"].map(lambda x: f"{x:.4f}")
    for col in display.columns[3:]:   # weight columns
        display[col] = display[col].map(lambda x: f"{x*100:.2f}%")

    for col_idx, col_name in enumerate(display.columns, 1):
        _header_style(ws.cell(row=1, column=col_idx))
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(display.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 3 – Stock Statistics
# ---------------------------------------------------------------------------

def _write_stats_sheet(wb, expected_returns, volatilities, daily_returns):
    ws = wb.create_sheet("Stock Statistics")

    tickers = expected_returns.index.tolist()
    corr    = daily_returns.corr()

    # Individual stats table
    for col_idx, h in enumerate(["Ticker", "Ann. Return", "Ann. Volatility"], 1):
        _header_style(ws.cell(row=1, column=col_idx))
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, ticker in enumerate(tickers, start=2):
        ws.cell(row=row_idx, column=1, value=ticker)
        ws.cell(row=row_idx, column=2, value=f"{expected_returns[ticker]*100:.2f}%")
        ws.cell(row=row_idx, column=3, value=f"{volatilities[ticker]*100:.2f}%")

    # Correlation matrix
    corr_start = len(tickers) + 4

    ws.cell(row=corr_start, column=1, value="Correlation Matrix").font = Font(bold=True, size=12)

    header_row = corr_start + 1
    ws.cell(row=header_row, column=1, value="")
    for col_idx, ticker in enumerate(tickers, 2):
        cell = ws.cell(row=header_row, column=col_idx, value=ticker)
        _header_style(cell, _ACCENT_FILL)
        cell.font = Font(bold=True, color="000000")

    for row_offset, row_ticker in enumerate(tickers):
        r = header_row + 1 + row_offset
        row_hdr = ws.cell(row=r, column=1, value=row_ticker)
        row_hdr.font = Font(bold=True)
        row_hdr.fill = PatternFill("solid", fgColor=_ACCENT_FILL)

        for col_offset, col_ticker in enumerate(tickers):
            val  = corr.loc[row_ticker, col_ticker]
            cell = ws.cell(row=r, column=2 + col_offset, value=round(val, 4))
            if row_ticker != col_ticker:
                if val > 0.7:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")   # green = highly correlated
                elif val < 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")   # red = negatively correlated

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 4 – Efficient Frontier Chart (embedded image)
# ---------------------------------------------------------------------------

def _write_chart_sheet(wb, chart_filename=CHART_FILENAME):
    from openpyxl.drawing.image import Image as XLImage

    ws = wb.create_sheet("Efficient Frontier Chart")
    ws.sheet_view.showGridLines = False

    if not os.path.exists(chart_filename):
        ws.cell(row=1, column=1, value=f"Chart file not found: {chart_filename}")
        return

    img = XLImage(chart_filename)
    img.width  = 960
    img.height = 640
    img.anchor = "B2"
    ws.add_image(img)

    # Size the sheet area to frame the image cleanly
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = int(img.width / 7.5) + 2
    ws.row_dimensions[1].height = 12


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_excel(
    optimal: dict,
    simulation: pd.DataFrame,
    expected_returns: pd.Series,
    volatilities: pd.Series,
    daily_returns: "pd.DataFrame",
    tickers: list[str] = TICKERS,
    filename: str = EXCEL_FILENAME,
    chart_filename: str = CHART_FILENAME,
) -> None:
    """Create the Excel workbook and write all four sheets."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    _write_optimal_sheet(wb, optimal, expected_returns, volatilities, tickers)
    _write_frontier_sheet(wb, simulation)
    _write_stats_sheet(wb, expected_returns, volatilities, daily_returns)
    _write_chart_sheet(wb, chart_filename)

    wb.save(filename)
    print(f"Excel workbook saved -> {filename}")
